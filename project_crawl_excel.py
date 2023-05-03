import requests
import re
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl

#보안 뉴스 크롤링
now=datetime.now().strftime("%Y-%m-%d")
url = 'http://www.boannews.com/media/t_list.asp'

headers = {
    'User-Agent': 'Mozilla/5.0',   
    'Content-Type': 'text/html; charset=utf-8'
}

# 엑셀파일로 저장
workbook = openpyxl.Workbook()#엑셀workbook
worksheet = workbook.active #엑셀파일이 열리면 sheet

worksheet['A1']="Titles"
worksheet['B1']="Contents"
worksheet['C1']="Writers"
worksheet['D1']="Dates"

row_index=2

for page in range(1,5):
    req = requests.get(f"{url}?Page={page}&kind=", headers=headers)
    soup=BeautifulSoup(req.text,"lxml")

    titles=soup.select("#news_area > div > a > span")
    contents=soup.select("#news_area > div > a.news_content")
    writers=soup.select("#news_area > div > span")

    #찍히는지 확인 # zip 활용해서 excel 저장
    for title, content, writer in zip(titles, contents, writers):
        print(f"기사 제목: {title.string},\n기사 내용: {content.string},\n기자: {writer.string}")
        print("-----------------------")

        worksheet.cell(row=row_index, column=1, value=title.string)
        worksheet.cell(row=row_index, column=2, value=content.string)
        worksheet.cell(row=row_index, column=3, value=writer.string)
        worksheet.cell(row=row_index, column=4, value=now)
        row_index += 1

    if not titles:
        break

    print(f"page:{page}")

workbook.save(f'boannews_{now}.xlsx')
        
